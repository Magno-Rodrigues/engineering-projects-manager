"""Reporting service for generating financial reports."""
import os
from datetime import datetime
from typing import List, Optional, Tuple

from app import db
from app.models.financial_report import FinancialReport, REPORT_TYPES
from app.models.project import Project
from app.services.financial_service import (
    FinancialTransactionService,
    FinancialBudgetService,
    FinancialEVMService,
)
from app.services.cash_flow_service import CashFlowService
from app.services.evm_analysis_service import EVMAnalysisService

REPORTS_DIR = '/tmp/reports'


class ReportingService:
    """Service class for generating and managing financial reports."""

    @staticmethod
    def generate_excel_report(
        project_id: int,
        report_type: str,
        generated_by: int = None,
    ) -> Tuple[Optional[FinancialReport], Optional[str]]:
        """Create Excel file with multiple sheets, record in FinancialReport, return report object."""
        try:
            import openpyxl
        except ImportError:
            return None, 'openpyxl is not installed.'

        if report_type not in REPORT_TYPES:
            return None, f'Invalid report type. Must be one of: {", ".join(REPORT_TYPES)}.'

        project = db.session.get(Project, project_id)
        if not project:
            return None, 'Project not found.'

        os.makedirs(REPORTS_DIR, exist_ok=True)
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f'report_{project_id}_{report_type}_{timestamp}.xlsx'
        file_path = os.path.join(REPORTS_DIR, filename)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        if report_type in ('executive', 'detailed'):
            ReportingService._add_summary_sheet(wb, project_id, project.name)
        if report_type == 'detailed':
            ReportingService._add_transactions_sheet(wb, project_id)
            ReportingService._add_budget_sheet(wb, project_id)
        if report_type == 'evm':
            ReportingService._add_evm_sheet(wb, project_id)
        if report_type == 'cash_flow':
            ReportingService._add_cash_flow_sheet(wb, project_id)

        wb.save(file_path)

        report = FinancialReport(
            project_id=project_id,
            report_type=report_type,
            generated_by=generated_by,
            file_path=file_path,
            format='xlsx',
            status='finalized',
        )
        db.session.add(report)
        db.session.commit()
        return report, None

    @staticmethod
    def _add_summary_sheet(wb, project_id: int, project_name: str):
        """Add a summary sheet with KPIs and EVM summary."""
        ws = wb.create_sheet('Summary')
        ws.append(['Financial Summary', project_name])
        ws.append([])
        summary = FinancialTransactionService.get_project_summary(project_id)
        ws.append(['Total Expenses', summary['total_expenses']])
        ws.append(['Total Revenues', summary['total_revenues']])
        ws.append(['Net Cash Flow', summary['net_cash_flow']])
        ws.append([])
        evm = EVMAnalysisService.get_evm_summary(project_id)
        if evm:
            ws.append(['EVM Summary'])
            ws.append(['CPI', evm['cpi']])
            ws.append(['SPI', evm['spi']])
            ws.append(['EAC', evm['eac']])
            ws.append(['ETC', evm['etc']])

    @staticmethod
    def _add_transactions_sheet(wb, project_id: int):
        """Add a transactions sheet."""
        ws = wb.create_sheet('Transactions')
        ws.append(['Date', 'Type', 'Category', 'Description', 'Amount', 'Status'])
        txns = FinancialTransactionService.get_project_transactions(project_id)
        for t in txns:
            ws.append([
                str(t.transaction_date), t.type, t.category,
                t.description, float(t.amount), t.payment_status,
            ])

    @staticmethod
    def _add_budget_sheet(wb, project_id: int):
        """Add a budget sheet."""
        ws = wb.create_sheet('Budget')
        ws.append(['Budget', 'Status', 'Total Planned', 'Baseline Date'])
        budgets = FinancialBudgetService.get_project_budgets(project_id)
        for b in budgets:
            ws.append([b.id, b.status, float(b.total_planned_budget), str(b.baseline_date)])

    @staticmethod
    def _add_evm_sheet(wb, project_id: int):
        """Add an EVM data sheet."""
        ws = wb.create_sheet('EVM')
        ws.append(['Date', 'BAC', 'AC', 'EV', 'PV', 'CPI', 'SPI'])
        data = EVMAnalysisService.get_scurve_data(project_id)
        trend = EVMAnalysisService.get_performance_trend(project_id)
        reports = FinancialEVMService.get_project_evm_reports(project_id)
        sorted_reports = sorted(reports, key=lambda r: r.report_date)
        for i, row in enumerate(data):
            cpi = trend[i]['cpi'] if i < len(trend) else ''
            spi = trend[i]['spi'] if i < len(trend) else ''
            bac = float(sorted_reports[i].bac) if i < len(sorted_reports) else ''
            ws.append([row['date'], bac, row['ac'], row['ev'], row['pv'], cpi, spi])

    @staticmethod
    def _add_cash_flow_sheet(wb, project_id: int):
        """Add a monthly cash flow sheet."""
        ws = wb.create_sheet('Cash Flow')
        ws.append(['Month', 'Inflows', 'Outflows', 'Net', 'Accumulated'])
        data = CashFlowService.get_monthly_cash_flow(project_id)
        for row in data:
            ws.append([
                row['month'], row['inflows'], row['outflows'],
                row['net_cash_flow'], row['accumulated_cash_flow'],
            ])

    @staticmethod
    def get_project_reports(project_id: int) -> List[FinancialReport]:
        """Return all FinancialReport records for a project."""
        return FinancialReport.query.filter_by(project_id=project_id).order_by(
            FinancialReport.generated_at.desc()
        ).all()

    @staticmethod
    def delete_report(report_id: int) -> Tuple[bool, Optional[str]]:
        """Delete a report record (and file if it exists)."""
        report = db.session.get(FinancialReport, report_id)
        if not report:
            return False, 'Report not found.'
        if report.file_path and os.path.exists(report.file_path):
            try:
                os.remove(report.file_path)
            except OSError:
                pass
        db.session.delete(report)
        db.session.commit()
        return True, None
