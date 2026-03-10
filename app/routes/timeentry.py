"""Time entry (apontamentos) routes."""
from datetime import datetime, date as date_type, timedelta
from flask import Blueprint, render_template, redirect, url_for, flash, request, Response, jsonify
from flask_login import login_required, current_user
from app.utils.decorators import admin_required
from app.services.timeentry_service import TimeEntryService
from app.services.project_service import ProjectService
from app.models.time_entry import HOUR_TYPES

timeentry_bp = Blueprint('timeentry', __name__, url_prefix='/apontamentos')


def _parse_date(date_str):
    """Parse a date string (YYYY-MM-DD) into a date object or return None."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _parse_int(value):
    """Parse a string to int or return None."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _normalize_hours(hours_str):
    """Normalize hours from HH:MM to HH:MM:SS by appending :00 if seconds are missing."""
    if hours_str and len(hours_str) == 5 and hours_str[2] == ':':
        return hours_str + ':00'
    return hours_str


def _build_calendar_grid(start_date, end_date, entered_dates):
    """Build a list-of-weeks grid for the pending dashboard calendar.

    Each cell is a dict with keys: date, status
    Status values: 'outside' | 'weekend' | 'completed' | 'future' | 'pending'
    """
    today = date_type.today()
    # Align to Monday of the first week
    monday = start_date - timedelta(days=start_date.weekday())
    # Align to Sunday of the last week
    sunday = end_date + timedelta(days=(6 - end_date.weekday()))

    weeks = []
    current = monday
    while current <= sunday:
        week = []
        for _ in range(7):
            if current < start_date or current > end_date:
                status = 'outside'
            elif current.weekday() >= 5:
                status = 'weekend'
            elif current in entered_dates:
                status = 'completed'
            elif current > today:
                status = 'future'
            else:
                status = 'pending'
            week.append({'date': current, 'status': status})
            current += timedelta(days=1)
        weeks.append(week)
    return weeks


# ── Measurement Cycle routes (admin only) ─────────────────────────────────────

@timeentry_bp.route('/ciclos/')
@login_required
@admin_required
def cycles_index():
    """List all measurement cycles (admin only)."""
    cycles = TimeEntryService.get_all_cycles()
    return render_template('apontamentos/ciclos/index.html', cycles=cycles)


@timeentry_bp.route('/ciclos/new', methods=['GET', 'POST'])
@login_required
@admin_required
def cycles_create():
    """Create a new measurement cycle (admin only)."""
    if request.method == 'POST':
        start_day = _parse_int(request.form.get('start_day'))
        start_date = _parse_date(request.form.get('start_date'))
        end_date = _parse_date(request.form.get('end_date'))
        is_active = bool(request.form.get('is_active'))

        cycle, error = TimeEntryService.create_cycle(
            start_day=start_day or 0,
            start_date=start_date,
            end_date=end_date,
            is_active=is_active,
            created_by=current_user.id,
        )
        if cycle:
            flash('Ciclo de medição criado com sucesso.', 'success')
            return redirect(url_for('timeentry.cycles_index'))
        flash(error, 'error')
    return render_template('apontamentos/ciclos/form.html', cycle=None)


@timeentry_bp.route('/ciclos/<int:cycle_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def cycles_edit(cycle_id: int):
    """Edit an existing measurement cycle (admin only)."""
    cycle = TimeEntryService.get_cycle(cycle_id)
    if not cycle:
        flash('Ciclo de medição não encontrado.', 'error')
        return redirect(url_for('timeentry.cycles_index'))

    if request.method == 'POST':
        start_day = _parse_int(request.form.get('start_day'))
        start_date = _parse_date(request.form.get('start_date'))
        end_date = _parse_date(request.form.get('end_date'))
        is_active = bool(request.form.get('is_active'))

        updated, error = TimeEntryService.update_cycle(cycle_id, {
            'start_day': start_day or cycle.start_day,
            'start_date': start_date or cycle.start_date,
            'end_date': end_date or cycle.end_date,
            'is_active': is_active,
        })
        if updated:
            flash('Ciclo de medição atualizado com sucesso.', 'success')
            return redirect(url_for('timeentry.cycles_index'))
        flash(error, 'error')
    return render_template('apontamentos/ciclos/form.html', cycle=cycle)


# ── Time Entry routes ──────────────────────────────────────────────────────────

@timeentry_bp.route('/')
@login_required
def index():
    """List time entries with pagination, summary, and filters."""
    is_admin = current_user.role == 'admin'
    project_id = _parse_int(request.args.get('project_id'))
    cycle_id = _parse_int(request.args.get('cycle_id'))
    work_date = _parse_date(request.args.get('work_date'))
    search = request.args.get('search', '').strip()
    user_id_filter = _parse_int(request.args.get('user_id')) if is_admin else None
    page = max(1, _parse_int(request.args.get('page')) or 1)
    per_page = 50

    entries_all = TimeEntryService.get_time_entries(
        user_id=current_user.id,
        is_admin=is_admin,
        project_id=project_id,
        cycle_id=cycle_id,
        work_date=work_date,
        search=search,
        filter_user_id=user_id_filter,
    )

    # Summary stats
    total_entries = len(entries_all)
    total_hours_seconds = 0
    last_entry_date = None
    for e in entries_all:
        if e.hours_worked:
            try:
                parts = str(e.hours_worked).split(':')
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0
                total_hours_seconds += h * 3600 + m * 60 + s
            except (ValueError, IndexError):
                pass
        if last_entry_date is None or (e.work_date and e.work_date > last_entry_date):
            last_entry_date = e.work_date

    total_hours_display = f"{total_hours_seconds // 3600:02d}:{(total_hours_seconds % 3600) // 60:02d}"
    avg_hours_display = ''
    if total_entries > 0:
        avg_s = total_hours_seconds // total_entries
        avg_hours_display = f"{avg_s // 3600:02d}:{(avg_s % 3600) // 60:02d}"

    # Pagination
    total_pages = max(1, (total_entries + per_page - 1) // per_page)
    page = min(page, total_pages)
    entries = entries_all[(page - 1) * per_page: page * per_page]

    projects = ProjectService.get_user_projects(current_user.id, include_all=True)
    cycles = TimeEntryService.get_all_cycles()
    active_cycle = TimeEntryService.get_active_cycle()

    users = []
    if is_admin:
        from app.models.user import User
        users = User.query.order_by(User.username).all()

    # ── Pending-dates dashboard data ───────────────────────────────────────
    pending_dashboard = None
    if active_cycle:
        all_weekdays, entered_dates, pending, future, completed = \
            TimeEntryService.get_pending_cycle_dates(current_user.id, active_cycle)
        total_days = len(all_weekdays)
        completed_days = len(completed)
        pending_days = len(pending)
        completion_pct = round(completed_days / total_days * 100) if total_days > 0 else 0
        calendar_grid = _build_calendar_grid(
            active_cycle.start_date, active_cycle.end_date, entered_dates
        )
        days_remaining = max(0, (active_cycle.end_date - date_type.today()).days)
        if days_remaining > 7:
            days_urgency = 'green'
        elif days_remaining >= 3:
            days_urgency = 'yellow'
        else:
            days_urgency = 'red'
        pending_dashboard = {
            'all_weekdays': all_weekdays,
            'entered_dates': entered_dates,
            'pending': pending,
            'future': future,
            'completed': completed,
            'total_days': total_days,
            'completed_days': completed_days,
            'pending_days': pending_days,
            'completion_pct': completion_pct,
            'calendar_grid': calendar_grid,
            'days_remaining': days_remaining,
            'days_urgency': days_urgency,
        }

    return render_template(
        'apontamentos/index.html',
        entries=entries,
        projects=projects,
        cycles=cycles,
        active_cycle=active_cycle,
        selected_project_id=project_id,
        selected_cycle_id=cycle_id,
        selected_work_date=request.args.get('work_date', ''),
        search=search,
        selected_user_id=user_id_filter,
        users=users,
        page=page,
        total_pages=total_pages,
        total_entries=total_entries,
        total_hours_display=total_hours_display,
        avg_hours_display=avg_hours_display,
        last_entry_date=last_entry_date,
        pending_dashboard=pending_dashboard,
    )


@timeentry_bp.route('/new', methods=['GET', 'POST'])
@login_required
def create():
    """Create a new time entry."""
    projects = ProjectService.get_user_projects(current_user.id, include_all=True, exclude_blocked=True)
    active_cycle = TimeEntryService.get_active_cycle()

    if request.method == 'POST':
        work_date = _parse_date(request.form.get('work_date'))
        entry, error = TimeEntryService.create_time_entry(
            project_id=_parse_int(request.form.get('project_id')),
            user_id=current_user.id,
            main_activity=request.form.get('main_activity'),
            work_date=work_date,
            hours_worked=_normalize_hours(request.form.get('hours_worked')),
            hour_type=request.form.get('hour_type'),
            discipline=request.form.get('discipline'),
            sub_activity=request.form.get('sub_activity'),
            observation=request.form.get('observation'),
        )
        if entry:
            flash('Apontamento criado com sucesso.', 'success')
            return redirect(url_for('timeentry.index'))
        flash(error, 'error')

    return render_template(
        'apontamentos/create.html',
        projects=projects,
        active_cycle=active_cycle,
        hour_types=HOUR_TYPES,
    )


@timeentry_bp.route('/<int:entry_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(entry_id: int):
    """Edit an existing time entry."""
    entry = TimeEntryService.get_time_entry(entry_id)
    if not entry:
        flash('Apontamento não encontrado.', 'error')
        return redirect(url_for('timeentry.index'))

    is_admin = current_user.role == 'admin'

    # Permission check before rendering
    if not is_admin and entry.user_id != current_user.id:
        flash('Você não tem permissão para editar este apontamento.', 'error')
        return redirect(url_for('timeentry.index'))

    projects = ProjectService.get_user_projects(current_user.id, include_all=True, exclude_blocked=True)
    active_cycle = TimeEntryService.get_active_cycle()

    # Block non-admin from editing entries whose work_date is outside the active cycle
    if not is_admin and (
        not active_cycle or
        not (active_cycle.start_date <= entry.work_date <= active_cycle.end_date)
    ):
        flash('Apenas o administrador pode editar apontamentos de ciclos anteriores.', 'error')
        return redirect(url_for('timeentry.index'))

    if request.method == 'POST':
        work_date = _parse_date(request.form.get('work_date'))
        data = {
            'project_id': _parse_int(request.form.get('project_id')),
            'discipline': request.form.get('discipline'),
            'main_activity': request.form.get('main_activity'),
            'sub_activity': request.form.get('sub_activity'),
            'work_date': work_date,
            'hours_worked': _normalize_hours(request.form.get('hours_worked')),
            'hour_type': request.form.get('hour_type'),
            'observation': request.form.get('observation'),
        }
        updated, error = TimeEntryService.update_time_entry(
            entry_id=entry_id,
            data=data,
            is_admin=is_admin,
            current_user_id=current_user.id,
        )
        if updated:
            flash('Apontamento atualizado com sucesso.', 'success')
            return redirect(url_for('timeentry.index'))
        flash(error, 'error')

    return render_template(
        'apontamentos/edit.html',
        entry=entry,
        projects=projects,
        active_cycle=active_cycle,
        hour_types=HOUR_TYPES,
    )


@timeentry_bp.route('/<int:entry_id>/delete', methods=['POST'])
@login_required
def delete(entry_id: int):
    """Delete a time entry."""
    is_admin = current_user.role == 'admin'
    success, error = TimeEntryService.delete_time_entry(
        entry_id=entry_id,
        is_admin=is_admin,
        current_user_id=current_user.id,
    )
    if success:
        flash('Apontamento excluído com sucesso.', 'success')
    else:
        flash(error, 'error')
    return redirect(url_for('timeentry.index'))


@timeentry_bp.route('/bulk', methods=['POST'])
@login_required
def create_bulk():
    """Create time entries for multiple dates in one submission."""
    active_cycle = TimeEntryService.get_active_cycle()
    if not active_cycle:
        flash('Não há ciclo de medição ativo. Contate o administrador.', 'error')
        return redirect(url_for('timeentry.create'))

    dates_str = request.form.get('dates', '').strip()
    dates = []
    for d in dates_str.split(','):
        parsed = _parse_date(d.strip())
        if parsed:
            dates.append(parsed)

    if not dates:
        flash('Nenhuma data válida selecionada.', 'error')
        return redirect(url_for('timeentry.create'))

    entries, errors = TimeEntryService.create_bulk_time_entries(
        dates=dates,
        project_id=_parse_int(request.form.get('project_id')),
        user_id=current_user.id,
        main_activity=request.form.get('main_activity', ''),
        hours_worked=_normalize_hours(request.form.get('hours_worked', '')),
        hour_type=request.form.get('hour_type', ''),
        discipline=request.form.get('discipline') or None,
        sub_activity=request.form.get('sub_activity') or None,
        observation=request.form.get('observation') or None,
    )

    if entries:
        flash(
            f'{len(entries)} apontamento(s) criado(s) com sucesso.',
            'success',
        )
    for error in errors:
        flash(error, 'error')

    return redirect(url_for('timeentry.index'))


@timeentry_bp.route('/export-excel')
@login_required
def export_excel():
    """Export filtered time entries to Excel (.xlsx)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    is_admin = current_user.role == 'admin'
    project_id = _parse_int(request.args.get('project_id'))
    cycle_id = _parse_int(request.args.get('cycle_id'))
    work_date = _parse_date(request.args.get('work_date'))
    search = request.args.get('search', '').strip()
    user_id_filter = _parse_int(request.args.get('user_id')) if is_admin else None

    entries = TimeEntryService.get_time_entries(
        user_id=current_user.id,
        is_admin=is_admin,
        project_id=project_id,
        cycle_id=cycle_id,
        work_date=work_date,
        search=search,
        filter_user_id=user_id_filter,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Apontamentos'

    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['Data', 'Projeto', 'Atividade Principal', 'Sub-Atividade', 'Disciplina', 'Horas', 'Tipo']
    if is_admin:
        headers.append('Usuário')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, entry in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=entry.work_date.strftime('%d/%m/%Y') if entry.work_date else '')
        ws.cell(row=row_idx, column=2, value=entry.project.name if entry.project else '')
        ws.cell(row=row_idx, column=3, value=entry.main_activity or '')
        ws.cell(row=row_idx, column=4, value=entry.sub_activity or '')
        ws.cell(row=row_idx, column=5, value=entry.discipline or '')
        ws.cell(row=row_idx, column=6, value=str(entry.hours_worked) if entry.hours_worked else '')
        ws.cell(row=row_idx, column=7, value=entry.hour_type or '')
        if is_admin:
            ws.cell(row=row_idx, column=8, value=entry.user.username if entry.user else '')

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=apontamentos.xlsx'},
    )
